import json
import os
import base64
import io
import re
import psycopg2
import openpyxl

# Маппинг: номер листа (1-based) -> product_id
SHEET_TO_PRODUCT = {
    1: "jacks-universal-single",
    2: "jacks-universal-double",
    3: "jacks-cargo-single",
    4: "jacks-cargo-double",
    5: "jacks-universal-lock",
    6: "jacks-crusher",
    7: "jacks-transformer",
    8: "jacks-steprise",
    9: "jacks-cargo-lock",
    10: "jacks-aluminum-spring",
    11: "jacks-aluminum-lock",
    12: "jacks-aluminum-hydraulic",
    13: "jacks-medium-single",
    14: "jacks-pulling",
    15: "jacks-low",
    16: "jacks-low-telescopic",
    17: "jacks-telescopic",
    18: "jacks-autonomous",
    19: "jacks-rack",
    20: "jacks-hollow",
    21: "cylinders-power",
    22: "cylinders-hydraulic-pin",
    23: "jacks-cargo-rolling",
    24: "jacks-cargo-double-floating",
    25: "jacks-accessories",
}

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text, flags=re.UNICODE)
    text = re.sub(r'[\s_-]+', '-', text)
    return text[:100]


def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def handler(event: dict, context) -> dict:
    """GET — данные модельного ряда из БД. POST — загрузка Excel, сохранение по листам."""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    method = event.get('httpMethod', 'GET')

    # GET: возвращаем данные по product_id
    if method == 'GET':
        product_id = (event.get('queryStringParameters') or {}).get('product_id', '')
        if not product_id:
            return {'statusCode': 400, 'headers': CORS, 'body': json.dumps({'error': 'product_id required'})}

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT cols, models FROM product_models WHERE product_id = %s", (product_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if not row:
            return {'statusCode': 200, 'headers': CORS, 'body': json.dumps({'cols': [], 'models': []})}

        return {'statusCode': 200, 'headers': CORS, 'body': json.dumps({'cols': row[0], 'models': row[1]})}

    # POST: загрузка Excel файла
    body = json.loads(event.get('body') or '{}')
    file_b64 = body.get('file')

    if not file_b64:
        return {'statusCode': 400, 'headers': CORS, 'body': json.dumps({'error': 'file required'})}

    file_bytes = base64.b64decode(file_b64)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    conn = get_conn()
    cur = conn.cursor()

    saved = []

    for sheet_index, sheet_name in enumerate(wb.sheetnames, 1):
        product_id = SHEET_TO_PRODUCT.get(sheet_index)
        if not product_id:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Ищем строку-заголовок: первая строка где >= 2 непустых ячеек
        header_row_idx = None
        for i, row in enumerate(rows):
            non_empty = [c for c in row if c is not None and str(c).strip() != '']
            if len(non_empty) >= 2:
                header_row_idx = i
                break

        if header_row_idx is None:
            continue

        raw_headers = rows[header_row_idx]
        headers = [str(c).strip() if c is not None else '' for c in raw_headers]
        cols = [{'key': f'col{i}', 'label': h} for i, h in enumerate(headers) if h]

        models = []
        for row in rows[header_row_idx + 1:]:
            non_empty = [c for c in row if c is not None and str(c).strip() != '']
            if not non_empty:
                continue
            obj = {'model': str(row[0]).strip() if row[0] is not None else ''}
            for i, h in enumerate(headers):
                if h:
                    val = row[i] if i < len(row) else None
                    obj[f'col{i}'] = str(val).strip() if val is not None else ''
            models.append(obj)

        cur.execute(
            """
            INSERT INTO product_models (product_id, cols, models, updated_at)
            VALUES (%s, %s, %s, NOW())
            ON CONFLICT (product_id) DO UPDATE
            SET cols = EXCLUDED.cols, models = EXCLUDED.models, updated_at = NOW()
            """,
            (product_id, json.dumps(cols, ensure_ascii=False), json.dumps(models, ensure_ascii=False)),
        )
        saved.append({'sheet': sheet_name, 'product_id': product_id, 'rows': len(models)})

    conn.commit()
    cur.close()
    conn.close()

    return {
        'statusCode': 200,
        'headers': CORS,
        'body': json.dumps({'ok': True, 'saved': saved}, ensure_ascii=False),
    }
